#!/usr/bin/env python3
"""
Canvas LMS Inactive Learner & Engagement Tracking Report Generator
==================================================================
This script automates the retrieval of student engagement, assignment submissions,
and weekly activity logs from one or multiple Canvas course shells, classifies students 
into actionable intervention categories, and exports a professionally designed Excel report.

Requirements:
- Python 3.7+
- Canvas API Access Token (with Teacher/Admin permissions)
- Active internet connection

Usage:
  python canvas_engagement_tracker.py
"""

# ==============================================================================
# CONFIGURATION - MANUAL UPDATES HERE
# ==============================================================================
# Load environment variables from .env if present
import os
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_env_path):
    with open(_env_path, "r") as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())

API_URL = os.environ.get("CANVAS_API_URL", "https://edgewood.instructure.com/api/v1")
ACCESS_TOKEN = os.environ.get("CANVAS_ACCESS_TOKEN") or os.environ.get("CANVAS_FALLBACK_TOKEN")
COURSE_IDENTIFIER = os.environ.get("COURSE_IDENTIFIER", "EDU-900-UAD4-Spring-2020")  # Can be single ID/Code or comma-separated list (e.g. "643, 1567")
COURSE_DURATION_WEEKS = int(os.environ.get("COURSE_DURATION_WEEKS", 6))

# ==============================================================================
# DEPENDENCY MANAGEMENT (SELF-INSTALLATION BOOTSTRAP)
# ==============================================================================
import os
import sys
import json
import time
import datetime
from datetime import timezone, timedelta
import urllib.parse
import traceback
import concurrent.futures

def install_and_import(package):
    try:
        __import__(package)
    except ImportError:
        print(f"[*] Package '{package}' not found. Attempting auto-installation...")
        import subprocess
        try:
            subprocess.run([sys.executable, "-m", "pip", "install", package], check=True)
            print(f"[+] Successfully installed '{package}'")
        except Exception as e:
            print(f"[-] Failed to install '{package}' automatically: {e}")
            print(f"[-] Please install it manually: pip install {package}")
            sys.exit(1)

# Ensure required libraries are installed
install_and_import("requests")
install_and_import("openpyxl")

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CANVAS API CLIENT
# ==============================================================================
class CanvasAPIClient:
    """A robust, rate-limit aware Canvas LMS API client."""
    def __init__(self, base_url, token):
        self.base_url = base_url.rstrip('/')
        self.headers = {"Authorization": f"Bearer {token}"}
        self.session = requests.Session()
        self.session.headers.update(self.headers)

    def request(self, method, endpoint, params=None, data=None):
        """Sends an HTTP request with automatic rate limiting back-off and error handling."""
        url = endpoint if endpoint.startswith('http') else f"{self.base_url}/{endpoint.lstrip('/')}"
        
        while True:
            try:
                response = self.session.request(method, url, params=params, json=data, timeout=30)
                
                # Canvas API rate limiting check (X-Rate-Limit-Remaining header)
                rate_limit = response.headers.get("X-Rate-Limit-Remaining")
                if rate_limit:
                    try:
                        remaining = float(rate_limit)
                        if remaining < 10.0:
                            # Sleep to let the rate limit bucket refill
                            sleep_time = 2.0 if remaining < 5.0 else 0.5
                            time.sleep(sleep_time)
                    except ValueError:
                        pass
                
                # Check for rate-limiting status code
                if response.status_code == 403 and "rate limit" in response.text.lower():
                    print("[!] Rate limit hit. Sleeping for 5 seconds...")
                    time.sleep(5)
                    continue

                response.raise_for_status()
                return response
            except requests.exceptions.HTTPError as e:
                # Handle specific HTTP error codes
                if e.response.status_code == 401:
                    raise Exception("Canvas API Auth Error: Invalid or expired Access Token.")
                elif e.response.status_code == 404:
                    raise Exception(f"Canvas Resource Not Found: {url}")
                else:
                    raise Exception(f"Canvas API HTTP Error {e.response.status_code}: {e.response.text}")
            except requests.exceptions.RequestException as e:
                raise Exception(f"Canvas API Network Error: {e}")

    def get_paginated(self, endpoint, params=None):
        """Fetches all items from a paginated API endpoint using Canvas Link header pagination."""
        if params is None:
            params = {}
        if "per_page" not in params:
            params["per_page"] = 100

        all_items = []
        response = self.request("GET", endpoint, params=params)
        
        try:
            items = response.json()
            if isinstance(items, list):
                all_items.extend(items)
            elif isinstance(items, dict):
                return items
        except ValueError:
            raise Exception("Invalid JSON response from Canvas API.")

        while True:
            link_header = response.headers.get("Link")
            if not link_header:
                break

            links = parse_link_header(link_header)
            next_url = links.get("next")
            if not next_url:
                break

            response = self.request("GET", next_url)
            try:
                items = response.json()
                if isinstance(items, list):
                    all_items.extend(items)
            except ValueError:
                raise Exception("Invalid JSON response from paginated Canvas API.")

        return all_items

def parse_link_header(link_header):
    """Parses the Canvas Link header to extract next, prev, first, last relations."""
    links = {}
    parts = link_header.split(',')
    for part in parts:
        section = part.split(';')
        if len(section) < 2:
            continue
        url = section[0].strip().lstrip('<').rstrip('>')
        for param in section[1:]:
            param = param.strip()
            if param.startswith('rel='):
                rel_name = param.split('=')[1].strip('"')
                links[rel_name] = url
                break
    return links

# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================
def parse_iso_datetime(dt_str):
    """Parses a Canvas ISO8601 datetime string to a timezone-aware UTC datetime."""
    if not dt_str:
        return None
    
    if dt_str.endswith('Z'):
        dt_str = dt_str[:-1] + '+00:00'
    
    try:
        dt = datetime.datetime.fromisoformat(dt_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except ValueError:
        for fmt in ('%Y-%m-%dT%H:%M:%S%z', '%Y-%m-%dT%H:%M:%S.%f%z', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                dt = datetime.datetime.strptime(dt_str, fmt)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt
            except ValueError:
                continue
    return None

def format_duration(seconds):
    """Converts seconds into a human-readable HH:MM:SS format."""
    if seconds is None:
        return "00:00:00"
    total_seconds = int(seconds)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    secs = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"

def format_datetime(dt):
    """Formats a datetime object to YYYY-MM-DD HH:MM:SS."""
    if not dt:
        return "-"
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def format_date(dt):
    """Formats a datetime object to YYYY-MM-DD."""
    if not dt:
        return "-"
    return dt.strftime("%Y-%m-%d")

# ==============================================================================
# DATA RETRIEVAL WORKFLOW
# ==============================================================================
def find_course(client, identifier):
    """Finds a course by its numeric ID, SIS course ID, or searches for it by name/code."""
    print(f"[*] Locating course shell for: '{identifier}'...")
    
    # 1. Try treating identifier as numeric database ID
    if str(identifier).isdigit():
        try:
            course = client.request("GET", f"courses/{identifier}").json()
            print(f"[+] Found course by Database ID: {course.get('name')}")
            return course
        except Exception:
            pass

    # 2. Try treating identifier as SIS ID (prefixed search)
    encoded_sis_id = urllib.parse.quote(f"sis_course_id:{identifier}")
    try:
        course = client.request("GET", f"courses/{encoded_sis_id}").json()
        print(f"[+] Found course by SIS Course ID: {course.get('name')}")
        return course
    except Exception:
        pass

    # 3. Try global search for course code / name
    try:
        courses = client.get_paginated("courses", {"search_term": identifier})
        for c in courses:
            if c.get("course_code") == identifier or c.get("sis_course_id") == identifier or c.get("name") == identifier:
                print(f"[+] Found course by search term matching code/name: {c.get('name')}")
                return c
        if courses:
            c = courses[0]
            print(f"[!] Warning: No exact code match. Using closest search result: {c.get('name')} (ID: {c.get('id')})")
            return c
    except Exception as e:
        print(f"[-] Search query failed: {e}")

    raise Exception(f"Failed to find course '{identifier}' in Canvas. Please check the Course ID or Code.")

def extract_course_details(client, course, duration_weeks):
    """Extracts summary details of a course."""
    course_id = course["id"]
    
    print(f"[*] [{course.get('course_code') or course_id}] Retrieving modules...")
    modules = client.get_paginated(f"courses/{course_id}/modules")
    active_modules = [m for m in modules if m.get("published") is not False]
    
    print(f"[*] [{course.get('course_code') or course_id}] Retrieving student enrollments...")
    enrollments = client.get_paginated(f"courses/{course_id}/enrollments", {"type[]": "StudentEnrollment"})
    student_ids = {e["user_id"] for e in enrollments if e.get("enrollment_state") in ["active", "invited", "completed"]}
    
    start_date = parse_iso_datetime(course.get("start_at"))
    end_date = parse_iso_datetime(course.get("end_at"))
    
    if not start_date:
        dates = [parse_iso_datetime(e.get("created_at")) for e in enrollments if e.get("created_at")]
        if dates:
            start_date = min(dates)
            print(f"[!] Start Date was empty. Using earliest student enrollment date: {format_date(start_date)}")
    if not start_date:
        start_date = parse_iso_datetime(course.get("created_at"))
        if start_date:
            print(f"[!] Start Date was empty. Using course creation date: {format_date(start_date)}")
    if not start_date:
        start_date = datetime.datetime.now(timezone.utc) - timedelta(weeks=duration_weeks)
        print(f"[!] Start Date could not be resolved. Defaulting to: {format_date(start_date)}")

    if not end_date:
        end_date = start_date + timedelta(weeks=duration_weeks)

    return {
        "lms_code": "Canvas LMS",
        "course_code": course.get("course_code", "-"),
        "course_name": course.get("name", "-"),
        "course_start_date": start_date,
        "course_end_date": end_date,
        "total_modules": len(active_modules),
        "total_learners_enrolled": len(student_ids)
    }

def fetch_sections_mapping(client, course_id):
    """Fetches course sections and maps section_id to section name (cohort)."""
    print(f"[*] [{course_id}] Retrieving course sections...")
    sections = client.get_paginated(f"courses/{course_id}/sections")
    sections_map = {s["id"]: s["name"] for s in sections}
    return sections_map

def fetch_course_data(client, course_id, course_code, sections_map, start_date, duration_weeks):
    """Fetches all learners, assignments, submissions, and detailed student activity concurrently."""
    print(f"[*] [{course_id}] Fetching student list...")
    enrollments = client.get_paginated(f"courses/{course_id}/enrollments", {"type[]": "StudentEnrollment"})
    
    # Fallback to resolve student IDs if they are missing (N/A) from the primary token's response
    has_missing_sis = any(not (env.get("sis_user_id") or (env.get("user") or {}).get("sis_user_id")) for env in enrollments)
    fallback_token = os.environ.get("CANVAS_FALLBACK_TOKEN")
    if has_missing_sis and fallback_token:
        print(f"[*] [{course_id}] Primary token returned missing SIS IDs. Attempting fallback token...")
        try:
            fallback_client = CanvasAPIClient(client.base_url, fallback_token)
            fallback_enrollments = fallback_client.get_paginated(f"courses/{course_id}/enrollments", {"type[]": "StudentEnrollment"})
            fallback_map = {fe.get("user_id"): fe for fe in fallback_enrollments if fe.get("user_id")}
            
            # Merge fallback data into the main enrollments
            for env in enrollments:
                u_id = env.get("user_id")
                if u_id in fallback_map:
                    fe = fallback_map[u_id]
                    if fe.get("sis_user_id"):
                        env["sis_user_id"] = fe["sis_user_id"]
                    if "user" in env and "user" in fe:
                        fe_user = fe["user"] or {}
                        env_user = env["user"] or {}
                        if fe_user.get("sis_user_id"):
                            env_user["sis_user_id"] = fe_user["sis_user_id"]
                        if fe_user.get("integration_id"):
                            env_user["integration_id"] = fe_user["integration_id"]
        except Exception as e:
            print(f"[!] Warning: Fallback token resolution failed: {e}")
            
    student_records = {}
    for env in enrollments:
        user = env.get("user")
        if not user:
            continue
        user_id = user["id"]
        
        section_id = env.get("course_section_id")
        
        # Extract cohort from course_code (e.g. EDU-900-UAD4-Spring-2020 -> UAD4)
        cohort_name = "-"
        if course_code:
            parts = course_code.split("-")
            if len(parts) >= 3:
                cohort_name = parts[2]
            else:
                cohort_name = sections_map.get(section_id, f"Section {section_id}")
        else:
            cohort_name = sections_map.get(section_id, f"Section {section_id}")
        
        last_activity_env = parse_iso_datetime(env.get("last_activity_at"))
        total_activity_time = env.get("total_activity_time", 0) or 0
        
        if user_id not in student_records:
            # Resolve Student ID following priority rules (filtering out email-like IDs):
            login_id_val = user.get("login_id") or env.get("login_id")
            if login_id_val and "@" in str(login_id_val):
                login_id_val = None

            resolved_student_id = (
                env.get("sis_user_id") or 
                user.get("sis_user_id") or 
                user.get("integration_id") or 
                login_id_val or 
                (str(user_id) if user_id is not None else None) or 
                "N/A"
            )
            student_records[user_id] = {
                "user_id": user_id,
                "student_id": resolved_student_id,
                "name": user.get("name", "Unknown Learner"),
                "email": user.get("email") or user.get("login_id") or "-",
                "status": env.get("enrollment_state", "active"),
                "cohorts": {cohort_name},
                "last_access_date": last_activity_env,
                "last_activity_timestamp": last_activity_env,
                "total_activity_time": total_activity_time,
                "raw_enrollments": [env]
            }
        else:
            student_records[user_id]["cohorts"].add(cohort_name)
            student_records[user_id]["raw_enrollments"].append(env)
            student_records[user_id]["total_activity_time"] += total_activity_time
            if last_activity_env:
                current_last = student_records[user_id]["last_activity_timestamp"]
                if not current_last or last_activity_env > current_last:
                    student_records[user_id]["last_activity_timestamp"] = last_activity_env
                    student_records[user_id]["last_access_date"] = last_activity_env

    print(f"[*] [{course_id}] Retrieving published course assignments...")
    assignments = client.get_paginated(f"courses/{course_id}/assignments")
    published_assignments = [a for a in assignments if a.get("published") is not False and a.get("omit_from_final_grade") is not True and a.get("grading_type") != "not_graded"]
    print(f"[+] [{course_id}] Found {len(published_assignments)} active assignments.")

    print(f"[*] [{course_id}] Retrieving learner submissions...")
    submissions_list = client.get_paginated(
        f"courses/{course_id}/students/submissions",
        {"student_ids[]": "all", "include[]": "assignment"}
    )
    
    submissions_by_student = {}
    for sub in submissions_list:
        student_id = sub.get("user_id")
        if student_id not in submissions_by_student:
            submissions_by_student[student_id] = []
        submissions_by_student[student_id].append(sub)

    current_time = datetime.datetime.now(timezone.utc)
    delta_now = current_time - start_date
    raw_course_week = (delta_now.days // 7) + 1
    current_course_week = min(max(1, raw_course_week), duration_weeks)
    
    if raw_course_week > duration_weeks:
        print(f"[*] [{course_id}] Course has completed. Evaluating based on final week: Week {current_course_week}")
    else:
        print(f"[*] [{course_id}] Current week in course is: Week {current_course_week}")

    student_weekly_activity = {}
    total_students = len(student_records)
    print(f"[*] [{course_id}] Fetching weekly activity analytics concurrently (15 threads)...")
    
    def fetch_single_student_activity(uid):
        try:
            activity_data = client.request(
                "GET", f"courses/{course_id}/analytics/users/{uid}/activity"
            ).json()
            return uid, activity_data
        except Exception:
            return uid, {"page_views": {}, "participations": []}

    with concurrent.futures.ThreadPoolExecutor(max_workers=15) as executor:
        future_to_uid = {
            executor.submit(fetch_single_student_activity, uid): uid
            for uid in student_records
        }
        
        completed = 0
        for future in concurrent.futures.as_completed(future_to_uid):
            uid = future_to_uid[future]
            try:
                user_id, activity_data = future.result()
                student_weekly_activity[user_id] = activity_data
            except Exception:
                student_weekly_activity[uid] = {"page_views": {}, "participations": []}
            completed += 1
            print(f"    Progress: {completed}/{total_students} analytics retrieved...", end="\r")
            
    print(f"\n[+] [{course_id}] Analytics retrieval complete.")
    return student_records, published_assignments, submissions_by_student, student_weekly_activity, current_course_week

# ==============================================================================
# REPORT CALCULATIONS
# ==============================================================================
def process_report_data(student_records, assignments, submissions_map, activity_map, course_start, total_weeks, current_course_week):
    """Computes metrics, weekly breakdowns, and simplified classification categories for students."""
    processed_students = []
    current_time = datetime.datetime.now(timezone.utc)
    
    week_ranges = []
    for w in range(1, total_weeks + 1):
        w_start = course_start + timedelta(weeks=w-1)
        w_end = course_start + timedelta(weeks=w) - timedelta(seconds=1)
        week_ranges.append((w, w_start, w_end))

    active_assignment_ids = {a["id"] for a in assignments}

    for user_id, student in student_records.items():
        student_subs = submissions_map.get(user_id, [])
        student_subs = [s for s in student_subs if s.get("assignment_id") in active_assignment_ids]

        total_assignments = len(assignments)
        submitted_count = 0
        missing_count = 0
        overdue_count = 0
        on_time_count = 0

        sub_dict = {s["assignment_id"]: s for s in student_subs}
        missed_names = []
        on_time_names = []

        for assign in assignments:
            assign_id = assign["id"]
            assign_name = assign.get("name", "Unknown Assignment")
            due_at = parse_iso_datetime(assign.get("due_at"))
            sub = sub_dict.get(assign_id)

            has_submitted = False
            submitted_at = None
            is_late = False

            if sub:
                submitted_at = parse_iso_datetime(sub.get("submitted_at"))
                workflow_state = sub.get("workflow_state", "unsubmitted")
                submission_type = sub.get("submission_type")
                is_missing = sub.get("missing", False) or False
                
                # Check if this assignment requires an online submission
                sub_types = assign.get("submission_types", [])
                is_online = any(t in ["online_upload", "online_text_entry", "online_url", "media_recording", "discussion_topic", "online_quiz", "external_tool"] for t in sub_types)
                
                if workflow_state not in ["unsubmitted", "unsubmitted_late"] and not is_missing:
                    if is_online:
                        # For online assignments, student must have submitted online
                        if (submitted_at is not None) or (submission_type is not None):
                            has_submitted = True
                            is_late = sub.get("late", False) or False
                    else:
                        # For offline assignments (like on_paper or none), they don't submit online.
                        # We count it as submitted if it has been graded/scored and is not marked as missing.
                        if (sub.get("score") is not None) or (submitted_at is not None):
                            has_submitted = True
                            is_late = sub.get("late", False) or False

            if due_at:
                due_passed = due_at < current_time
                if has_submitted:
                    submitted_count += 1
                    if is_late or (submitted_at and submitted_at > due_at):
                        pass
                    else:
                        on_time_count += 1
                        on_time_names.append(assign_name)
                else:
                    if due_passed:
                        missing_count += 1
                        overdue_count += 1
                        missed_names.append(assign_name)
            else:
                if has_submitted:
                    submitted_count += 1
                    on_time_count += 1
                    on_time_names.append(assign_name)

        activity = activity_map.get(user_id, {"page_views": {}, "participations": []})
        page_views = activity.get("page_views", {})
        participations = activity.get("participations", [])

        parsed_views = []
        for pv_time_str, count in page_views.items():
            pv_time = parse_iso_datetime(pv_time_str)
            if pv_time:
                parsed_views.append((pv_time, count))

        parsed_participations = []
        for part in participations:
            part_time = parse_iso_datetime(part.get("created_at"))
            if part_time:
                parsed_participations.append(part_time)

        weekly_stats = {}
        total_course_activities = 0
        weekly_activities_counts = {}

        for w, w_start, w_end in week_ranges:
            w_views = sum(count for pv_time, count in parsed_views if w_start <= pv_time <= w_end)
            w_parts = sum(1 for part_time in parsed_participations if w_start <= part_time <= w_end)
            w_act_count = w_views + w_parts
            weekly_activities_counts[w] = w_act_count
            total_course_activities += w_act_count

            w_activity_times = []
            for pv_time, count in parsed_views:
                if count > 0 and w_start <= pv_time <= w_end:
                    w_activity_times.append(pv_time)
            for part_time in parsed_participations:
                if w_start <= part_time <= w_end:
                    w_activity_times.append(part_time)

            w_last_activity = max(w_activity_times) if w_activity_times else None
            weekly_stats[w] = {
                "start": w_start,
                "end": w_end,
                "activities": w_act_count,
                "last_activity": w_last_activity,
                "time_spent": 0
            }

        total_time_seconds = student.get("total_activity_time", 0) or 0
        if total_time_seconds > 0 and total_course_activities > 0:
            for w in range(1, total_weeks + 1):
                share = weekly_activities_counts[w] / total_course_activities
                weekly_stats[w]["time_spent"] = total_time_seconds * share
        elif total_time_seconds > 0:
            active_weeks_count = sum(1 for w, w_start, _ in week_ranges if current_time >= w_start)
            if active_weeks_count > 0:
                allocated = total_time_seconds / active_weeks_count
                for w, w_start, _ in week_ranges:
                    if current_time >= w_start:
                        weekly_stats[w]["time_spent"] = allocated

        all_activity_times = [pv_time for pv_time, count in parsed_views if count > 0] + parsed_participations
        overall_last_activity = max(all_activity_times) if all_activity_times else None
        canvas_last = student.get("last_activity_timestamp")
        if canvas_last:
            if not overall_last_activity or canvas_last > overall_last_activity:
                overall_last_activity = canvas_last

        engagement_hours = total_time_seconds / 3600.0
        has_logged_in = (overall_last_activity is not None) and (total_time_seconds > 0)

        # Calculate Weeks 1-4 combined engagement
        w1_to_w4_seconds = sum(
            weekly_stats[w]["time_spent"]
            for w in range(1, 5)
            if w in weekly_stats and isinstance(weekly_stats[w]["time_spent"], (int, float))
        )
        w1_to_w4_hours = w1_to_w4_seconds / 3600.0

        # Classification Logic based on user's exact simplified criteria
        if not has_logged_in or (total_course_activities == 0 and total_time_seconds == 0):
            category = "NO ACTIVITY"
            comment = "Learner has not logged into the platform. There is no activity."
        
        elif overdue_count > 0:
            category = "MISSED SUBMISSION"
            comment = "Learner has past due assignments or unsubmitted."
            
        elif engagement_hours < 1.0 and overdue_count == 0 and (submitted_count == on_time_count or total_assignments == 0):
            category = "LOW ACTIVITY - ASSIGNMENTS COMPLETED"
            comment = "Engagement is less than 1 hour but assignments completed on time."
            
        elif total_weeks >= 4 and current_course_week >= 4 and w1_to_w4_hours < 1.0:
            category = "NOT ACTIVE"
            comment = "Total combined engagement of Week 1 to Week 4 is less than 1 hour (applicable in or after Week 4)."
            
        else:
            category = "ACTIVE"
            comment = "Good engagement and on track."

        overall_activity = "Active" if category in ["ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED"] else "Inactive"

        cohorts_str = ", ".join(sorted(list(student["cohorts"])))
        student_data = {
            "cohort": cohorts_str,
            "student_id": student["student_id"],
            "name": student["name"],
            "email": student["email"],
            "status": student["status"],
            "last_activity_timestamp": overall_last_activity.replace(tzinfo=None) if overall_last_activity else None,
            "total_engagement_seconds": total_time_seconds,
            "total_assignments": total_assignments,
            "submitted_assignments": submitted_count,
            "missing_assignments": missing_count,
            "overdue_assignments": overdue_count,
            "on_time_submissions": on_time_count,
            "overall_activity": overall_activity,
            "category": category,
            "comment": comment,
            "weekly_data": {}
        }

        for w, w_start, w_end in week_ranges:
            if current_time < w_start:
                student_data["weekly_data"][w] = {
                    "time_spent": "-",
                    "status": "-"
                }
            else:
                w_stat = weekly_stats[w]
                w_time = w_stat["time_spent"]
                w_status = "MET" if w_time > 0 else "NOT MET"
                student_data["weekly_data"][w] = {
                    "time_spent": w_time,
                    "status": w_status
                }

        processed_students.append(student_data)
        
    return processed_students

# ==============================================================================
# EXCEL GENERATOR
# ==============================================================================
def create_excel_report(course_results, output_filename, duration_weeks=None):
    """
    Generates a professionally styled Excel report with:
    - Tab 1: Dashboard (Unified Course Summary + Guidelines)
    - Tab 2+: Detail tab per course
    """
    if duration_weeks is None:
        try:
            duration_weeks = len(course_results[0]["processed_students"][0]["weekly_data"])
        except (IndexError, KeyError, TypeError):
            duration_weeks = COURSE_DURATION_WEEKS

    print(f"[*] Generating combined Excel report: '{output_filename}'...")
    wb = openpyxl.Workbook()
    
    PRIMARY_COLOR = "1E293B"    # Dark Slate Header
    SECONDARY_COLOR = "475569"  # Secondary Header
    ZEBRA_COLOR = "F8FAFC"      # Zebra alternate
    BORDER_COLOR = "E2E8F0"     # Borders
    
    CATEGORY_STYLES = {
        "ACTIVE": {"fill": "DCFCE7", "font": "166534"},
        "LOW ACTIVITY - ASSIGNMENTS COMPLETED": {"fill": "E0F2FE", "font": "075985"},
        "MISSED SUBMISSION": {"fill": "FEF3C7", "font": "92400E"},
        "NOT ACTIVE": {"fill": "FEE2E2", "font": "991B1B"},
        "NO ACTIVITY": {"fill": "F1F5F9", "font": "475569"}
    }

    WEEKLY_STATUS_STYLES = {
        "MET": {"fill": "E8F5E9", "font": "2E7D32"},
        "NOT MET": {"fill": "FFEBEE", "font": "C62828"},
        "-": {"fill": "FFFFFF", "font": "000000"}
    }

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    double_bottom_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='double', color="000000")
    )

    def clean_sheet_name(name):
        for char in r":\/?*[]":
            name = name.replace(char, "")
        return name.strip()[:30]

    # --------------------------------------------------------------------------
    # TAB 1: COMBINED DASHBOARD
    # --------------------------------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True

    ws_dash["A1"] = "Canvas LMS Learner Engagement Portfolio Dashboard"
    ws_dash["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
    ws_dash["A2"] = f"Report Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_dash["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="777777")
    
    ws_dash["A4"] = "Portfolio Overview"
    ws_dash["A4"].font = Font(name="Segoe UI", size=13, bold=True, color="1E293B")
    
    headers_dash = [
        "Course Code", "Course Name", "Total Enrolled", 
        "ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED", "MISSED SUBMISSION", "NOT ACTIVE", "NO ACTIVITY"
    ]
    
    for c_idx, col_name in enumerate(headers_dash, 1):
        cell = ws_dash.cell(row=5, column=c_idx, value=col_name)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if c_idx > 2 else "left", vertical="center")
        cell.border = thin_border
        
    curr_row = 6
    total_portfolio_enrolled = 0
    total_portfolio_cats = {cat: 0 for cat in CATEGORY_STYLES}
    
    for res in course_results:
        details = res["course_details"]
        students = res["processed_students"]
        
        cat_counts = {cat: 0 for cat in CATEGORY_STYLES}
        for s in students:
            cat = s["category"]
            if cat in cat_counts:
                cat_counts[cat] += 1
                
        total_enrolled = len(students)
        total_portfolio_enrolled += total_enrolled
        for cat in CATEGORY_STYLES:
            total_portfolio_cats[cat] += cat_counts[cat]
            
        ws_dash.cell(row=curr_row, column=1, value=details["course_code"]).font = Font(name="Segoe UI", size=10, bold=True)
        ws_dash.cell(row=curr_row, column=1).border = thin_border
        
        ws_dash.cell(row=curr_row, column=2, value=details["course_name"]).font = Font(name="Segoe UI", size=10)
        ws_dash.cell(row=curr_row, column=2).border = thin_border
        
        ws_dash.cell(row=curr_row, column=3, value=total_enrolled).alignment = Alignment(horizontal="center")
        ws_dash.cell(row=curr_row, column=3).border = thin_border
        
        c_offset = 4
        for cat in CATEGORY_STYLES:
            count = cat_counts[cat]
            cell_cnt = ws_dash.cell(row=curr_row, column=c_offset, value=count)
            cell_cnt.alignment = Alignment(horizontal="center")
            cell_cnt.border = thin_border
            
            style = CATEGORY_STYLES[cat]
            cell_cnt.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
            cell_cnt.font = Font(name="Segoe UI", size=10, bold=count > 0, color=style["font"])
            c_offset += 1
            
        curr_row += 1
        
    # Totals Row
    ws_dash.cell(row=curr_row, column=1, value="Portfolio Total").font = Font(name="Segoe UI", size=10, bold=True)
    ws_dash.cell(row=curr_row, column=1).border = double_bottom_border
    ws_dash.cell(row=curr_row, column=1).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    ws_dash.cell(row=curr_row, column=2, value="").border = double_bottom_border
    ws_dash.cell(row=curr_row, column=2).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    t_enrolled_cell = ws_dash.cell(row=curr_row, column=3, value=total_portfolio_enrolled)
    t_enrolled_cell.font = Font(name="Segoe UI", size=10, bold=True)
    t_enrolled_cell.alignment = Alignment(horizontal="center")
    t_enrolled_cell.border = double_bottom_border
    t_enrolled_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    c_offset = 4
    for cat in CATEGORY_STYLES:
        t_count = total_portfolio_cats[cat]
        t_cell_cnt = ws_dash.cell(row=curr_row, column=c_offset, value=t_count)
        t_cell_cnt.font = Font(name="Segoe UI", size=10, bold=True)
        t_cell_cnt.alignment = Alignment(horizontal="center")
        t_cell_cnt.border = double_bottom_border
        t_cell_cnt.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        c_offset += 1
        
    curr_row += 3
    
    # 2. Stacked Metadata Cards
    ws_dash.cell(row=curr_row, column=1, value="Course Details Overview").font = Font(name="Segoe UI", size=13, bold=True, color="1E293B")
    curr_row += 1
    
    for res in course_results:
        details = res["course_details"]
        ws_dash.cell(row=curr_row, column=1, value=f"{details['course_name']} ({details['course_code']})").font = Font(name="Segoe UI", size=11, bold=True, color="475569")
        ws_dash.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=3)
        curr_row += 1
        
        course_meta_keys = [
            "LMS Platform", "Course Code", "Course Name", 
            "Course Start Date", "Course End Date", "Total Modules", "Total Learners Enrolled"
        ]
        course_meta_vals = [
            details["lms_code"],
            details["course_code"],
            details["course_name"],
            format_date(details["course_start_date"]),
            format_date(details["course_end_date"]),
            details["total_modules"],
            details["total_learners_enrolled"]
        ]
        
        for k, v in zip(course_meta_keys, course_meta_vals):
            ws_dash.cell(row=curr_row, column=1, value=k).font = Font(name="Segoe UI", size=10, bold=True, color="475569")
            ws_dash.cell(row=curr_row, column=1).fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            ws_dash.cell(row=curr_row, column=1).border = thin_border
            
            val_cell = ws_dash.cell(row=curr_row, column=2, value=v)
            val_cell.font = Font(name="Segoe UI", size=10)
            val_cell.border = thin_border
            val_cell.alignment = Alignment(horizontal="left")
            ws_dash.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
            ws_dash.cell(row=curr_row, column=3).border = thin_border
            curr_row += 1
            
        curr_row += 1
        
    curr_row += 1
    
    # 3. Category Criteria Guidelines Table
    ws_dash.cell(row=curr_row, column=1, value="Engagement Classification Criteria").font = Font(name="Segoe UI", size=13, bold=True, color="1E293B")
    curr_row += 1
    
    guidelines = [
        ("NO ACTIVITY", "Learner has not logged into the platform. There is no activity."),
        ("MISSED SUBMISSION", "Learner has past due assignments or unsubmitted."),
        ("NOT ACTIVE", "Total combined engagement of Week 1 to Week 4 is less than 1 hour (applicable in or after Week 4)."),
        ("LOW ACTIVITY - ASSIGNMENTS COMPLETED", "Engagement is less than 1 hour but assignments completed on time."),
        ("ACTIVE", "Good engagement and on track.")
    ]
    
    for cat_name, desc in guidelines:
        lbl_cell = ws_dash.cell(row=curr_row, column=1, value=cat_name)
        lbl_cell.font = Font(name="Segoe UI", size=9, bold=True)
        if cat_name in CATEGORY_STYLES:
            style = CATEGORY_STYLES[cat_name]
            lbl_cell.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
            lbl_cell.font = Font(name="Segoe UI", size=9, bold=True, color=style["font"])
        lbl_cell.border = thin_border
        
        desc_cell = ws_dash.cell(row=curr_row, column=2, value=desc)
        desc_cell.font = Font(name="Segoe UI", size=9)
        desc_cell.border = thin_border
        ws_dash.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=8)
        
        for c in range(3, 9):
            ws_dash.cell(row=curr_row, column=c).border = thin_border
            
        curr_row += 1

    # Adjust dashboard column widths
    for col in ws_dash.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_dash.column_dimensions[col_letter].width = max(max_len + 3, 12)
    ws_dash.column_dimensions['A'].width = 28
    ws_dash.column_dimensions['B'].width = 30

    # --------------------------------------------------------------------------
    # TAB 2+: COURSE DETAILS TABS
    # --------------------------------------------------------------------------
    for res in course_results:
        course_details = res["course_details"]
        processed_students = res["processed_students"]
        
        sheet_title = clean_sheet_name(f"Details-{course_details['course_code']}")
        ws_detail = wb.create_sheet(title=sheet_title)
        ws_detail.views.sheetView[0].showGridLines = True
        ws_detail.freeze_panes = "D3"

        # Headers Setup
        headers_row1 = [
            ("Learner Information", 4),
            ("Activity Summary", 2),
            ("Assignment Tracking", 5),
            ("Engagement Classification", 3),
        ]
        
        for w in range(1, duration_weeks + 1):
            w_start = course_details["course_start_date"] + timedelta(weeks=w-1)
            w_end = course_details["course_start_date"] + timedelta(weeks=w) - timedelta(seconds=1)
            date_str = f"Week {w} ({format_date(w_start)} to {format_date(w_end)})"
            headers_row1.append((date_str, 2))
            
        headers_row2 = [
            "Cohort", "Learner Name", "Official Email ID", "Enrollment Status",
            "Last Activity Timestamp", "Total Time Spent (HH:MM:SS)",
            "Total Assignments", "Submitted Assignments", "Missing Assignments", "Overdue Assignments", "On-Time Submissions",
            "Engagement Category", "Learner's Overall Activity", "Automated Comments"
        ]
        
        for w in range(1, duration_weeks + 1):
            headers_row2.extend([
                f"W{w} Duration (HH:MM:SS)",
                f"W{w} Target Status"
            ])

        # Write Group Headers (Row 1)
        col_idx = 1
        for group_name, span in headers_row1:
            start_col = col_idx
            end_col = col_idx + span - 1
            ws_detail.cell(row=1, column=start_col, value=group_name)
            ws_detail.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            
            cell = ws_detail.cell(row=1, column=start_col)
            cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for c in range(start_col, end_col + 1):
                ws_detail.cell(row=1, column=c).fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
                ws_detail.cell(row=1, column=c).border = thin_border
            col_idx += span

        # Write Column Headers (Row 2)
        for c_idx, col_name in enumerate(headers_row2, 1):
            cell = ws_detail.cell(row=2, column=c_idx, value=col_name)
            cell.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=SECONDARY_COLOR, end_color=SECONDARY_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Write Student Rows (Row 3+)
        start_row = 3
        for s_idx, student in enumerate(processed_students):
            curr_s_row = start_row + s_idx
            row_fill = PatternFill(start_color=ZEBRA_COLOR, end_color=ZEBRA_COLOR, fill_type="solid") if s_idx % 2 == 1 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Info fields
            ws_detail.cell(row=curr_s_row, column=1, value=student["cohort"]).alignment = Alignment(horizontal="left")
            ws_detail.cell(row=curr_s_row, column=2, value=student["name"]).alignment = Alignment(horizontal="left")
            ws_detail.cell(row=curr_s_row, column=3, value=student["email"]).alignment = Alignment(horizontal="left")
            ws_detail.cell(row=curr_s_row, column=4, value=(student["status"] or "active").capitalize()).alignment = Alignment(horizontal="center")
            
            # Last Activity
            last_act_val = student["last_activity_timestamp"]
            if isinstance(last_act_val, datetime.datetime) and last_act_val.tzinfo is not None:
                last_act_val = last_act_val.replace(tzinfo=None)
            last_act_cell = ws_detail.cell(row=curr_s_row, column=5, value=last_act_val)
            last_act_cell.alignment = Alignment(horizontal="center")
            if isinstance(last_act_val, datetime.datetime):
                last_act_cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                
            # Time Spent
            ws_detail.cell(row=curr_s_row, column=6, value=format_duration(student["total_engagement_seconds"])).alignment = Alignment(horizontal="center")
            
            # Assignment columns
            ws_detail.cell(row=curr_s_row, column=7, value=student["total_assignments"]).alignment = Alignment(horizontal="center")
            ws_detail.cell(row=curr_s_row, column=8, value=student["submitted_assignments"]).alignment = Alignment(horizontal="center")
            ws_detail.cell(row=curr_s_row, column=9, value=student["missing_assignments"]).alignment = Alignment(horizontal="center")
            ws_detail.cell(row=curr_s_row, column=10, value=student["overdue_assignments"]).alignment = Alignment(horizontal="center")
            ws_detail.cell(row=curr_s_row, column=11, value=student["on_time_submissions"]).alignment = Alignment(horizontal="center")
            
            # Engagement category
            cat_cell = ws_detail.cell(row=curr_s_row, column=12, value=student["category"])
            cat_cell.alignment = Alignment(horizontal="center")
            if student["category"] in CATEGORY_STYLES:
                style = CATEGORY_STYLES[student["category"]]
                cat_cell.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
                cat_cell.font = Font(name="Segoe UI", size=10, bold=True, color=style["font"])
                
            # Overall Activity
            oa_cell = ws_detail.cell(row=curr_s_row, column=13, value=student["overall_activity"])
            oa_cell.alignment = Alignment(horizontal="center")
            if student["overall_activity"] == "Active":
                oa_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                oa_cell.font = Font(name="Segoe UI", size=10, bold=True, color="166534")
            else:
                oa_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                oa_cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
            
            # Comments
            ws_detail.cell(row=curr_s_row, column=14, value=student["comment"]).alignment = Alignment(horizontal="left")
            
            # Apply basic layout borders/fonts to student info cells
            for c in range(1, 15):
                cell = ws_detail.cell(row=curr_s_row, column=c)
                cell.border = thin_border
                if c != 14:  # Keep specific fill/font for Category cell
                    cell.fill = row_fill
                    cell.font = Font(name="Segoe UI", size=10)

            # Weekly Columns
            c_offset = 15
            for w in range(1, duration_weeks + 1):
                w_data = student["weekly_data"].get(w, {"time_spent": "-", "status": "-"})
                w_time = w_data["time_spent"]
                w_status = w_data["status"]
                
                # Duration
                dur_val = format_duration(w_time) if isinstance(w_time, (int, float)) else w_time
                dur_cell = ws_detail.cell(row=curr_s_row, column=c_offset, value=dur_val)
                dur_cell.alignment = Alignment(horizontal="center")
                dur_cell.border = thin_border
                dur_cell.fill = row_fill
                dur_cell.font = Font(name="Segoe UI", size=10)
                
                # Status
                status_cell = ws_detail.cell(row=curr_s_row, column=c_offset + 1, value=w_status)
                status_cell.alignment = Alignment(horizontal="center")
                status_cell.border = thin_border
                
                if w_status in WEEKLY_STATUS_STYLES:
                    w_style = WEEKLY_STATUS_STYLES[w_status]
                    status_cell.fill = PatternFill(start_color=w_style["fill"], end_color=w_style["fill"], fill_type="solid")
                    status_cell.font = Font(name="Segoe UI", size=10, bold=True, color=w_style["font"])
                else:
                    status_cell.fill = row_fill
                    status_cell.font = Font(name="Segoe UI", size=10)
                    
                c_offset += 2

        # Auto-fit columns
        for col in ws_detail.columns:
            max_len = 0
            for cell in col:
                if cell.row == 1:
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws_detail.column_dimensions[col_letter].width = max(max_len + 4, 10)

        # Custom layouts
        ws_detail.column_dimensions['A'].width = 16  # Cohort
        ws_detail.column_dimensions['B'].width = 22  # Name
        ws_detail.column_dimensions['C'].width = 26  # Email
        ws_detail.column_dimensions['D'].width = 18  # Enrollment Status
        ws_detail.column_dimensions['E'].width = 22  # Last Activity Timestamp
        ws_detail.column_dimensions['F'].width = 24  # Total Time Spent
        ws_detail.column_dimensions['L'].width = 25  # Category
        ws_detail.column_dimensions['M'].width = 25  # Overall Activity
        ws_detail.column_dimensions['N'].width = 45  # Comments

        # Enable Auto-Filter on detail sheet
        last_col_letter = get_column_letter(col_idx - 1)
        ws_detail.auto_filter.ref = f"A2:{last_col_letter}{len(processed_students) + start_row - 1}"

    # --------------------------------------------------------------------------
    # TAB N: CONSOLIDATED DETAILS
    # --------------------------------------------------------------------------
    ws_con = wb.create_sheet(title="Consolidated Details")
    ws_con.views.sheetView[0].showGridLines = True
    ws_con.freeze_panes = "E3"

    headers_con_row1 = [
        ("Course Details", 1),
        ("Learner Information", 4),
        ("Activity Summary", 2),
        ("Assignment Tracking", 5),
        ("Engagement Classification", 3),
    ]
    for w in range(1, duration_weeks + 1):
        headers_con_row1.append((f"Week {w}", 2))

    headers_con_row2 = [
        "Course Name",
        "Cohort", "Learner Name", "Official Email ID", "Enrollment Status",
        "Last Activity Timestamp", "Total Time Spent (HH:MM:SS)",
        "Total Assignments", "Submitted Assignments", "Missing Assignments", "Overdue Assignments", "On-Time Submissions",
        "Engagement Category", "Learner's Overall Activity", "Automated Comments"
    ]
    for w in range(1, duration_weeks + 1):
        headers_con_row2.extend([
            f"W{w} Duration (HH:MM:SS)",
            f"W{w} Target Status"
        ])

    # Write Group Headers (Row 1)
    col_idx = 1
    for group_name, span in headers_con_row1:
        start_col = col_idx
        end_col = col_idx + span - 1
        ws_con.cell(row=1, column=start_col, value=group_name)
        ws_con.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        
        cell = ws_con.cell(row=1, column=start_col)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for c in range(start_col, end_col + 1):
            ws_con.cell(row=1, column=c).fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
            ws_con.cell(row=1, column=c).border = thin_border
        col_idx += span

    # Write Column Headers (Row 2)
    for c_idx, col_name in enumerate(headers_con_row2, 1):
        cell = ws_con.cell(row=2, column=c_idx, value=col_name)
        cell.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=SECONDARY_COLOR, end_color=SECONDARY_COLOR, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Write Student Rows (Row 3+)
    start_row = 3
    global_s_idx = 0
    for res in course_results:
        course_details = res["course_details"]
        processed_students = res["processed_students"]
        course_name = course_details["course_name"]
        
        for student in processed_students:
            curr_s_row = start_row + global_s_idx
            row_fill = PatternFill(start_color=ZEBRA_COLOR, end_color=ZEBRA_COLOR, fill_type="solid") if global_s_idx % 2 == 1 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Course Name
            ws_con.cell(row=curr_s_row, column=1, value=course_name).alignment = Alignment(horizontal="left")
            
            # Info fields
            ws_con.cell(row=curr_s_row, column=2, value=student["cohort"]).alignment = Alignment(horizontal="left")
            ws_con.cell(row=curr_s_row, column=3, value=student["name"]).alignment = Alignment(horizontal="left")
            ws_con.cell(row=curr_s_row, column=4, value=student["email"]).alignment = Alignment(horizontal="left")
            ws_con.cell(row=curr_s_row, column=5, value=(student["status"] or "active").capitalize()).alignment = Alignment(horizontal="center")
            
            # Last Activity
            last_act_val = student["last_activity_timestamp"]
            if isinstance(last_act_val, datetime.datetime) and last_act_val.tzinfo is not None:
                last_act_val = last_act_val.replace(tzinfo=None)
            last_act_cell = ws_con.cell(row=curr_s_row, column=6, value=last_act_val)
            last_act_cell.alignment = Alignment(horizontal="center")
            if isinstance(last_act_val, datetime.datetime):
                last_act_cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                
            # Time Spent
            ws_con.cell(row=curr_s_row, column=7, value=format_duration(student["total_engagement_seconds"])).alignment = Alignment(horizontal="center")
            
            # Assignment columns
            ws_con.cell(row=curr_s_row, column=8, value=student["total_assignments"]).alignment = Alignment(horizontal="center")
            ws_con.cell(row=curr_s_row, column=9, value=student["submitted_assignments"]).alignment = Alignment(horizontal="center")
            ws_con.cell(row=curr_s_row, column=10, value=student["missing_assignments"]).alignment = Alignment(horizontal="center")
            ws_con.cell(row=curr_s_row, column=11, value=student["overdue_assignments"]).alignment = Alignment(horizontal="center")
            ws_con.cell(row=curr_s_row, column=12, value=student["on_time_submissions"]).alignment = Alignment(horizontal="center")
            
            # Engagement category
            cat_cell = ws_con.cell(row=curr_s_row, column=13, value=student["category"])
            cat_cell.alignment = Alignment(horizontal="center")
            if student["category"] in CATEGORY_STYLES:
                style = CATEGORY_STYLES[student["category"]]
                cat_cell.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
                cat_cell.font = Font(name="Segoe UI", size=10, bold=True, color=style["font"])
                
            # Overall Activity
            oa_cell = ws_con.cell(row=curr_s_row, column=14, value=student["overall_activity"])
            oa_cell.alignment = Alignment(horizontal="center")
            if student["overall_activity"] == "Active":
                oa_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                oa_cell.font = Font(name="Segoe UI", size=10, bold=True, color="166534")
            else:
                oa_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                oa_cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
            
            # Comments
            ws_con.cell(row=curr_s_row, column=15, value=student["comment"]).alignment = Alignment(horizontal="left")
            
            # Apply basic layout borders/fonts to student info cells
            for c in range(1, 16):
                cell = ws_con.cell(row=curr_s_row, column=c)
                cell.border = thin_border
                if c != 15:  # Keep specific fill/font for Category cell
                    cell.fill = row_fill
                    cell.font = Font(name="Segoe UI", size=10)

            # Weekly Columns
            c_offset = 16
            for w in range(1, duration_weeks + 1):
                w_data = student["weekly_data"].get(w, {"time_spent": "-", "status": "-"})
                w_time = w_data["time_spent"]
                w_status = w_data["status"]
                
                # Duration
                dur_val = format_duration(w_time) if isinstance(w_time, (int, float)) else w_time
                dur_cell = ws_con.cell(row=curr_s_row, column=c_offset, value=dur_val)
                dur_cell.alignment = Alignment(horizontal="center")
                dur_cell.border = thin_border
                dur_cell.fill = row_fill
                dur_cell.font = Font(name="Segoe UI", size=10)
                
                # Status
                status_cell = ws_con.cell(row=curr_s_row, column=c_offset + 1, value=w_status)
                status_cell.alignment = Alignment(horizontal="center")
                status_cell.border = thin_border
                
                if w_status in WEEKLY_STATUS_STYLES:
                    w_style = WEEKLY_STATUS_STYLES[w_status]
                    status_cell.fill = PatternFill(start_color=w_style["fill"], end_color=w_style["fill"], fill_type="solid")
                    status_cell.font = Font(name="Segoe UI", size=10, bold=True, color=w_style["font"])
                else:
                    status_cell.fill = row_fill
                    status_cell.font = Font(name="Segoe UI", size=10)
                    
                c_offset += 2
                
            global_s_idx += 1

    # Auto-fit columns for Consolidated Details
    for col in ws_con.columns:
        max_len = 0
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws_con.column_dimensions[col_letter].width = max(max_len + 4, 10)

    # Custom layouts for Consolidated Details
    ws_con.column_dimensions['A'].width = 28  # Course Name
    ws_con.column_dimensions['B'].width = 16  # Cohort
    ws_con.column_dimensions['C'].width = 22  # Name
    ws_con.column_dimensions['D'].width = 26  # Email
    ws_con.column_dimensions['E'].width = 18  # Enrollment Status
    ws_con.column_dimensions['F'].width = 22  # Last Activity Timestamp
    ws_con.column_dimensions['G'].width = 24  # Total Time Spent
    ws_con.column_dimensions['M'].width = 25  # Category
    ws_con.column_dimensions['N'].width = 25  # Overall Activity
    ws_con.column_dimensions['O'].width = 45  # Comments

    # Enable Auto-Filter on consolidated sheet
    last_col_letter = get_column_letter(col_idx - 1)
    ws_con.auto_filter.ref = f"A2:{last_col_letter}{global_s_idx + start_row - 1}"

    # Save Workbook with lock handling
    try:
        wb.save(output_filename)
        print(f"[+] Saved successfully to: {output_filename}")
        return output_filename
    except PermissionError:
        print(f"[!] Warning: Permission Denied while saving to '{output_filename}'.")
        print("    This usually means the spreadsheet is currently open in Excel.")
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        name_parts = os.path.splitext(output_filename)
        fallback_filename = f"{name_parts[0]}_{timestamp}{name_parts[1]}"
        
        print(f"[*] Saving to fallback location instead: {fallback_filename}")
        try:
            wb.save(fallback_filename)
            print(f"[+] Saved successfully to fallback file: {fallback_filename}")
            return fallback_filename
        except Exception as fallback_err:
            raise Exception(f"Failed to save to both primary and fallback locations. Error: {fallback_err}")

# ==============================================================================
# SAMPLE / MOCK DATA RUNNER FOR DEMO & TESTING
# ==============================================================================
def generate_mock_data(duration_weeks=6):
    """Generates synthetic data representing two mock courses for local verification."""
    print(f"[*] Generating mock learner and engagement data for 2 courses ({duration_weeks} weeks)...")
    
    # --------------------------------------------------------------------------
    # COURSE 1 (Spring 2026)
    # --------------------------------------------------------------------------
    weeks_passed = max(1, duration_weeks - 2)
    c1_details = {
        "lms_code": "Canvas LMS (Mock)",
        "course_code": "EDU-900-UAD4-Spring-2020",
        "course_name": "Instructional Design Essentials",
        "course_start_date": datetime.datetime.now(timezone.utc) - timedelta(weeks=weeks_passed, days=3),
        "course_end_date": datetime.datetime.now(timezone.utc) + timedelta(weeks=1, days=4),
        "total_modules": 8,
        "total_learners_enrolled": 5
    }

    c1_students = [
        # Student 1: Active, high engagement, on time
        {
            "cohort": "Section A",
            "student_id": "STU1001",
            "name": "Jane Miller",
            "email": "jane.miller@edgewood.edu",
            "status": "active",
            "last_activity_timestamp": datetime.datetime.now(timezone.utc) - timedelta(hours=2),
            "total_engagement_seconds": 25200, # 7 hours
            "total_assignments": 6,
            "submitted_assignments": 4,
            "missing_assignments": 0,
            "overdue_assignments": 0,
            "on_time_submissions": 4,
            "missed_names": "None",
            "on_time_names": "None",
            "category": "ACTIVE",
            "overall_activity": "Active" if "ACTIVE" in ["ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED"] else "Inactive",
            "comment": "Good engagement and on track.",
            "weekly_data": {}
        },
        # Student 2: Never logged in
        {
            "cohort": "Section A",
            "student_id": "STU1002",
            "name": "John Doe",
            "email": "john.doe@edgewood.edu",
            "status": "active",
            "last_activity_timestamp": None,
            "total_engagement_seconds": 0,
            "total_assignments": 6,
            "submitted_assignments": 0,
            "missing_assignments": 4,
            "overdue_assignments": 4,
            "on_time_submissions": 0,
            "missed_names": "None",
            "on_time_names": "None",
            "category": "NO ACTIVITY",
            "overall_activity": "Active" if "NO ACTIVITY" in ["ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED"] else "Inactive",
            "comment": "Learner has not logged into the platform. There is no activity.",
            "weekly_data": {}
        },
        # Student 3: Missed submission
        {
            "cohort": "Section B",
            "student_id": "STU1003",
            "name": "Robert Smith",
            "email": "robert.smith@edgewood.edu",
            "status": "active",
            "last_activity_timestamp": datetime.datetime.now(timezone.utc) - timedelta(days=1),
            "total_engagement_seconds": 18000, # 5 hours
            "total_assignments": 6,
            "submitted_assignments": 3,
            "missing_assignments": 1,
            "overdue_assignments": 1,
            "on_time_submissions": 3,
            "missed_names": "None",
            "on_time_names": "None",
            "category": "MISSED SUBMISSION",
            "overall_activity": "Active" if "MISSED SUBMISSION" in ["ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED"] else "Inactive",
            "comment": "Learner has past due assignments or unsubmitted.",
            "weekly_data": {}
        },
        # Student 4: Low activity, completed on time
        {
            "cohort": "Section B",
            "student_id": "STU1004",
            "name": "Emily Davis",
            "email": "emily.davis@edgewood.edu",
            "status": "active",
            "last_activity_timestamp": datetime.datetime.now(timezone.utc) - timedelta(days=3),
            "total_engagement_seconds": 2400, # 40 mins
            "total_assignments": 6,
            "submitted_assignments": 4,
            "missing_assignments": 0,
            "overdue_assignments": 0,
            "on_time_submissions": 4,
            "missed_names": "None",
            "on_time_names": "None",
            "category": "LOW ACTIVITY - ASSIGNMENTS COMPLETED",
            "overall_activity": "Active" if "LOW ACTIVITY - ASSIGNMENTS COMPLETED" in ["ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED"] else "Inactive",
            "comment": "Engagement is less than 1 hour but assignments completed on time.",
            "weekly_data": {}
        },
        # Student 5: Not active (Week 5, engagement under 1 hour, late submissions)
        {
            "cohort": "Section A",
            "student_id": "STU1005",
            "name": "Michael Brown",
            "email": "michael.brown@edgewood.edu",
            "status": "active",
            "last_activity_timestamp": datetime.datetime.now(timezone.utc) - timedelta(days=5),
            "total_engagement_seconds": 1500, # 25 mins
            "total_assignments": 6,
            "submitted_assignments": 4,
            "missing_assignments": 0,
            "overdue_assignments": 0,
            "on_time_submissions": 2,
            "missed_names": "None",
            "on_time_names": "None", # 2 late submissions
            "category": "NOT ACTIVE",
            "overall_activity": "Active" if "NOT ACTIVE" in ["ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED"] else "Inactive",
            "comment": "Total combined engagement of Week 1 to Week 4 is less than 1 hour (applicable in or after Week 4).",
            "weekly_data": {}
        }
    ]

    # Calculate weekly values for Course 1
    current_time = datetime.datetime.now(timezone.utc)
    for std in c1_students:
        for w in range(1, duration_weeks + 1):
            w_start = c1_details["course_start_date"] + timedelta(weeks=w-1)
            if current_time < w_start:
                std["weekly_data"][w] = {"time_spent": "-", "status": "-"}
            else:
                if std["category"] == "NO ACTIVITY":
                    std["weekly_data"][w] = {"time_spent": 0, "status": "NOT MET"}
                elif std["category"] == "ACTIVE":
                    std["weekly_data"][w] = {"time_spent": 5040, "status": "MET"}
                elif std["category"] == "MISSED SUBMISSION":
                    std["weekly_data"][w] = {"time_spent": 3600, "status": "MET"}
                elif std["category"] == "LOW ACTIVITY - ASSIGNMENTS COMPLETED":
                    std["weekly_data"][w] = {"time_spent": 480, "status": "MET"}
                else: # NOT ACTIVE
                    std["weekly_data"][w] = {"time_spent": 300 if w < 5 else 0, "status": "MET" if w < 5 else "NOT MET"}

    # --------------------------------------------------------------------------
    # COURSE 2 (Fall 2025)
    # --------------------------------------------------------------------------
    weeks_passed_c2 = max(1, duration_weeks - 1)
    c2_details = {
        "lms_code": "Canvas LMS (Mock)",
        "course_code": "BUS-903-UC9-Fall-2025",
        "course_name": "Analyzing Behavior in Organizations",
        "course_start_date": datetime.datetime.now(timezone.utc) - timedelta(weeks=weeks_passed_c2, days=2),
        "course_end_date": datetime.datetime.now(timezone.utc) + timedelta(days=5),
        "total_modules": 12,
        "total_learners_enrolled": 3
    }

    c2_students = [
        # Student 1: Active
        {
            "cohort": "Cohort 1",
            "student_id": "STU2001",
            "name": "Sarah Connor",
            "email": "sarah.connor@edgewood.edu",
            "status": "active",
            "last_activity_timestamp": datetime.datetime.now(timezone.utc) - timedelta(minutes=45),
            "total_engagement_seconds": 43200, # 12 hours
            "total_assignments": 8,
            "submitted_assignments": 7,
            "missing_assignments": 0,
            "overdue_assignments": 0,
            "on_time_submissions": 7,
            "missed_names": "None",
            "on_time_names": "None",
            "category": "ACTIVE",
            "overall_activity": "Active" if "ACTIVE" in ["ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED"] else "Inactive",
            "comment": "Good engagement and on track.",
            "weekly_data": {}
        },
        # Student 2: Missed Submission
        {
            "cohort": "Cohort 1",
            "student_id": "STU2002",
            "name": "John Connor",
            "email": "john.connor@edgewood.edu",
            "status": "active",
            "last_activity_timestamp": datetime.datetime.now(timezone.utc) - timedelta(days=2),
            "total_engagement_seconds": 14400, # 4 hours
            "total_assignments": 8,
            "submitted_assignments": 5,
            "missing_assignments": 2,
            "overdue_assignments": 2,
            "on_time_submissions": 5,
            "missed_names": "None",
            "on_time_names": "None",
            "category": "MISSED SUBMISSION",
            "overall_activity": "Active" if "MISSED SUBMISSION" in ["ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED"] else "Inactive",
            "comment": "Learner has past due assignments or unsubmitted.",
            "weekly_data": {}
        },
        # Student 3: No activity
        {
            "cohort": "Cohort 2",
            "student_id": "STU2003",
            "name": "T-800 Terminator",
            "email": "t800@cyberdyne.com",
            "status": "active",
            "last_activity_timestamp": None,
            "total_engagement_seconds": 0,
            "total_assignments": 8,
            "submitted_assignments": 0,
            "missing_assignments": 7,
            "overdue_assignments": 7,
            "on_time_submissions": 0,
            "missed_names": "None",
            "on_time_names": "None",
            "category": "NO ACTIVITY",
            "overall_activity": "Active" if "NO ACTIVITY" in ["ACTIVE", "LOW ACTIVITY - ASSIGNMENTS COMPLETED"] else "Inactive",
            "comment": "Learner has not logged into the platform. There is no activity.",
            "weekly_data": {}
        }
    ]

    # Calculate weekly values for Course 2
    for std in c2_students:
        for w in range(1, duration_weeks + 1):
            w_start = c2_details["course_start_date"] + timedelta(weeks=w-1)
            if current_time < w_start:
                std["weekly_data"][w] = {"time_spent": "-", "status": "-"}
            else:
                if std["category"] == "NO ACTIVITY":
                    std["weekly_data"][w] = {"time_spent": 0, "status": "NOT MET"}
                elif std["category"] == "ACTIVE":
                    std["weekly_data"][w] = {"time_spent": 7200, "status": "MET"}
                else: # MISSED SUBMISSION
                    std["weekly_data"][w] = {"time_spent": 2880, "status": "MET"}

    return [
        {"course_details": c1_details, "processed_students": c1_students},
        {"course_details": c2_details, "processed_students": c2_students}
    ]

# ==============================================================================
# MAIN DRIVER
# ==============================================================================
def main():
    import argparse
    parser = argparse.ArgumentParser(description="Generate Combined Canvas LMS Inactive Learner & Engagement Report.")
    parser.add_argument("--token", default=ACCESS_TOKEN, help="Canvas API Access Token")
    parser.add_argument("--course", default=COURSE_IDENTIFIER, help="Canvas Course ID(s) or SIS Code(s), comma-separated")
    parser.add_argument("--url", default=API_URL, help="Canvas LMS API Base URL")
    parser.add_argument("--weeks", type=int, default=COURSE_DURATION_WEEKS, help="Course duration in weeks")
    parser.add_argument("--mock", action="store_true", help="Generate report with mock data (offline test)")
    parser.add_argument("--output", default="Canvas_Engagement_Report_Combined.xlsx", help="Output Excel filename")
    args = parser.parse_args()

    print("====================================================================")
    print("      Canvas LMS Inactive Learner & Engagement Report Generator      ")
    print("====================================================================")
    
    if args.mock:
        print("[*] Running in OFFLINE MOCK MODE...")
        course_results = generate_mock_data(args.weeks)
        saved_filename = create_excel_report(course_results, args.output, args.weeks)
        print(f"[+] Demo report successfully generated: {saved_filename}")
        print("====================================================================")
        return

    # Check configurations
    if not args.token or args.token == "YOUR_CANVAS_API_ACCESS_TOKEN_HERE" or args.token.startswith("YOUR_"):
        print("[-] Error: Canvas API Access Token is not set.")
        print("    Please configure ACCESS_TOKEN at the top of the script or pass it via --token")
        sys.exit(1)
        
    course_input = args.course
    if not course_input or course_input == "EDU-900-UAD4-Spring-2020" or course_input == "YOUR_COURSE_SHELL_ID_OR_CODE_HERE":
        print("\n[?] Course shell ID or code is required.")
        try:
            course_input = input("    Please enter Canvas Course ID(s)/Code(s) (comma-separated if multiple, e.g. 643, 1567): ").strip()
        except KeyboardInterrupt:
            print("\n[-] Cancelled by user.")
            sys.exit(0)
        if not course_input:
            print("[-] Error: Course identifier cannot be empty.")
            sys.exit(1)

    # Parse comma-separated course list
    course_identifiers = [c.strip() for c in course_input.split(",") if c.strip()]
    if not course_identifiers:
        print("[-] Error: No valid course identifiers found.")
        sys.exit(1)

    print(f"[*] Connecting to Canvas LMS at: {args.url}")
    client = CanvasAPIClient(args.url, args.token)
    
    course_results = []
    
    try:
        for idx, cid in enumerate(course_identifiers, 1):
            print(f"\n[*] Processing Course {idx}/{len(course_identifiers)}: '{cid}'...")
            
            # 1. Locate Course Shell
            course = find_course(client, cid)
            
            # 2. Extract Course Metadata
            course_details = extract_course_details(client, course, args.weeks)
            print("--------------------------------------------------------------------")
            print(f"LMS Platform:   {course_details['lms_code']}")
            print(f"Course Code:    {course_details['course_code']}")
            print(f"Course Name:    {course_details['course_name']}")
            print(f"Start Date:     {format_date(course_details['course_start_date'])}")
            print(f"End Date:       {format_date(course_details['course_end_date'])}")
            print(f"Total Modules:  {course_details['total_modules']}")
            print(f"Enrolled Learners: {course_details['total_learners_enrolled']}")
            print("--------------------------------------------------------------------")
            
            if course_details['total_learners_enrolled'] == 0:
                print(f"[!] Warning: No enrolled students found for course '{cid}'. Skipping...")
                continue
                
            # 3. Retrieve Sections (Cohorts)
            sections_map = fetch_sections_mapping(client, course["id"])
            
            # 4. Fetch Course Activity & Submission Data
            student_records, assignments, submissions_map, activity_map, current_week = fetch_course_data(
                client, course["id"], course_details["course_code"], sections_map, course_details["course_start_date"], args.weeks
            )
            
            # 5. Process Engagement Data and Categorization
            processed_students = process_report_data(
                student_records, assignments, submissions_map, activity_map, 
                course_details["course_start_date"], args.weeks, current_week
            )
            
            course_results.append({
                "course_details": course_details,
                "processed_students": processed_students
            })
            
        if not course_results:
            print("[-] Error: No course results were retrieved or processed.")
            sys.exit(1)

        # 6. Generate Excel Report
        saved_filename = create_excel_report(course_results, args.output, args.weeks)
        print(f"\n[+] Process complete. Combined report saved as: {saved_filename}")
        print("====================================================================")

    except Exception as e:
        print(f"\n[-] Error encountered during execution:")
        print(f"    {e}")
        print("\n[*] Detailed Traceback:")
        traceback.print_exc()
        print("\n[!] Would you like to run in mock mode to check report styling?")
        print("    Run: python canvas_engagement_tracker.py --mock")
        print("====================================================================")
        sys.exit(1)

if __name__ == "__main__":
    main()
